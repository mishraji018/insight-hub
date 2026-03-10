import io
import matplotlib.pyplot as plt
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils import timezone
from analytics.models import SalesData, Prediction, Anomaly
from django.db.models import Sum, Avg

def generate_pdf_report(date_from: str, date_to: str, user_role: str) -> io.BytesIO:
    """Generates a PDF report with KPI summary and charts."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Data Retrieval
    sales_qs = SalesData.objects.filter(date__range=[date_from, date_to])
    predictions_qs = Prediction.objects.filter(date__range=[date_from, date_to])
    anomalies_qs = Anomaly.objects.filter(date__range=[date_from, date_to])

    # Header
    elements.append(Paragraph("Predictive Analytics Dashboard - Sales Report", styles['Title']))
    elements.append(Paragraph(f"Period: {date_from} to {date_to}", styles['Normal']))
    elements.append(Paragraph(f"Generated at: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # KPI Summary
    total_sales = sales_qs.aggregate(Sum('sales_amount'))['sales_amount__sum'] or 0
    avg_customers = sales_qs.aggregate(Avg('customers'))['customers__avg'] or 0
    
    kpi_data = [
        ["Metric", "Value"],
        ["Total Sales", f"${total_sales:,.2f}"],
        ["Avg. Customers", f"{avg_customers:,.0f}"],
        ["Anomalies Detected", f"{anomalies_qs.count()}"],
    ]
    
    t = Table(kpi_data, colWidths=[150, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(Paragraph("KPI Summary", styles['Heading2']))
    elements.append(t)
    elements.append(Spacer(1, 20))

    # Sales Trend Chart
    if not sales_qs.exists():
        elements.append(Paragraph("No sales data available for this range.", styles['Normal']))
    else:
        df = pd.DataFrame(list(sales_qs.values('date', 'sales_amount')))
        df = df.groupby('date')['sales_amount'].sum().reset_index()
        
        plt.figure(figsize=(6, 4))
        plt.plot(df['date'], df['sales_amount'], marker='o', linestyle='-', color='teal')
        plt.title('Daily Sales Volume')
        plt.xlabel('Date')
        plt.ylabel('Sales ($)')
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png')
        img_buffer.seek(0)
        plt.close()
        
        elements.append(Paragraph("Sales Trends", styles['Heading2']))
        elements.append(Image(img_buffer, width=400, height=250))
        elements.append(Spacer(1, 20))

    # Predictions Table
    if predictions_qs.exists():
        elements.append(Paragraph("AI Sales Forecasts (7-Day Sample)", styles['Heading2']))
        pred_data = [["Date", "Predicted Sales", "Confidence"]]
        for p in predictions_qs[:7]:
            pred_data.append([str(p.date), f"${p.predicted_sales:,.2f}", f"{p.confidence_score*100:.1f}%"])
        
        pt = Table(pred_data, colWidths=[100, 120, 100])
        pt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkcyan),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(pt)

    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_excel_report(date_from: str, date_to: str) -> io.BytesIO:
    """Generates a multi-sheet Excel report."""
    wb = Workbook()
    
    # Sheet 1: Raw Sales
    ws1 = wb.active
    ws1.title = "Sales Data"
    sales_qs = SalesData.objects.filter(date__range=[date_from, date_to]).values()
    
    if sales_qs:
        headers = list(sales_qs[0].keys())
        ws1.append(headers)
        for row in sales_qs:
            ws1.append(list(row.values()))
    
    # Sheet 2: Predictions
    ws2 = wb.create_sheet(title="Predictions")
    pred_qs = Prediction.objects.filter(date__range=[date_from, date_to]).values()
    if pred_qs:
        p_headers = list(pred_qs[0].keys())
        ws2.append(p_headers)
        for row in pred_qs:
            ws2.append(list(row.values()))

    # Sheet 3: Summary
    ws3 = wb.create_sheet(title="Summary")
    agg_sales = SalesData.objects.filter(date__range=[date_from, date_to])
    total = agg_sales.aggregate(Sum('sales_amount'))['sales_amount__sum'] or 0
    avg = agg_sales.aggregate(Avg('sales_amount'))['sales_amount__avg'] or 0
    
    ws3['A1'] = "Report Summary"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3.append([]) # Empty row
    ws3.append(["Metric", "Value"])
    ws3.append(["Total Revenue", total])
    ws3.append(["Avg Daily Sales", avg])
    ws3.append(["Generated At", str(timezone.now())])

    # Formatting
    for ws in [ws1, ws2, ws3]:
        for cell in ws[1]: # Header row
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        ws.freeze_panes = "A2"
        # Auto-width hack
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
